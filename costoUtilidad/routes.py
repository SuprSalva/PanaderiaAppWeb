from . import costoUtilidad
from flask import render_template, request, redirect, url_for, flash, session, jsonify, Response, current_app
from flask_login import login_required, current_user
from models import db
from utils.db_roles import get_role_engine
from functools import wraps
from auth import roles_required
import forms
import datetime
import io


def _call_sp(sp_name, params=()):
    conn = get_role_engine().raw_connection()
    try:
        cursor = conn.cursor()
        cursor.callproc(sp_name, params)
        cols = [d[0] for d in cursor.description]
        rows = [dict(zip(cols, row)) for row in cursor.fetchall()]
        cursor.close()
        return rows
    finally:
        conn.close()


def _call_sp_one(sp_name, params=()):
    rows = _call_sp(sp_name, params)
    return rows[0] if rows else None


def _call_sp_multi(sp_name, params=()):
    """Llama un SP y retorna todos los result-sets como lista de listas."""
    conn = get_role_engine().raw_connection()
    try:
        cursor = conn.cursor()
        cursor.callproc(sp_name, params)
        sets = []
        while True:
            if cursor.description:
                cols = [d[0] for d in cursor.description]
                sets.append([dict(zip(cols, r)) for r in cursor.fetchall()])
            else:
                sets.append([])
            if not cursor.nextset():
                break
        cursor.close()
        return sets
    finally:
        conn.close()


def _parse_decimal(valor):
    try:
        v = str(valor).strip()
        return float(v) if v else None
    except (ValueError, TypeError):
        return None


def _parse_date(valor, default=None):
    """Convierte string 'YYYY-MM-DD' a date; retorna default si falla."""
    try:
        return datetime.date.fromisoformat(str(valor).strip())
    except Exception:
        return default


def _serial(obj):
    import decimal
    if isinstance(obj, decimal.Decimal):
        return float(obj)
    if isinstance(obj, (datetime.date, datetime.datetime, datetime.time)):
        return str(obj)
    return str(obj)



@costoUtilidad.route('/costo-utilidad', methods=['GET'])
@login_required
@roles_required('admin')
def index_costo_utilidad():
    current_app.logger.info('Vista de Costo y Utilidad accesada | usuario: %s | fecha: %s', current_user.username, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    form   = forms.CostoUtilidadFiltroForm(request.args)
    buscar = request.args.get('buscar', '').strip()
    orden  = request.args.get('orden', 'nombre_asc')

    util_min = _parse_decimal(request.args.get('utilidad_min', ''))
    util_max = _parse_decimal(request.args.get('utilidad_max', ''))

    orden_sp_map = {
        'nombre_asc':  '',
        'margen_asc':  'margen_asc',
        'margen_desc': 'margen_desc',
        'costo_asc':   'costo_asc',
        'costo_desc':  'costo_desc',
    }
    orden_sp = orden_sp_map.get(orden, '')

    _kpis_default = {
        'total_productos':       0,
        'margen_prom':           None,
        'costo_prom':            None,
        'precio_prom':           None,
        'productos_margen_bajo': 0,
    }

    try:
        productos = _call_sp(
            'sp_reporte_costo_utilidad',
            (buscar or None, orden_sp or None, util_min, util_max)
        )
        kpis_raw = _call_sp_one('sp_kpi_costo_utilidad', ())
        kpis = {**_kpis_default, **(kpis_raw or {})}
    except Exception as e:
        current_app.logger.error('Error al consultar datos de costo utilidad | usuario: %s | error: %s | fecha: %s', current_user.username, str(e), datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        flash(f'Error al obtener datos: {e}', 'error')
        productos = []
        kpis = _kpis_default

    return render_template(
        'costoUtilidad/costoUtilidad.html',
        form=form,
        productos=productos,
        kpis=kpis,
        buscar=buscar,
        orden=orden,
        util_min=util_min,
        util_max=util_max,
    )


@costoUtilidad.route('/costo-utilidad/exportar-excel', methods=['GET'])
@login_required
@roles_required('admin')
def exportar_excel_costo():
    current_app.logger.info('Reporte excel Costo Utilidad solicitado | usuario: %s | fecha: %s', current_user.username, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        current_app.logger.error('Fallo al exportar excel costo utilidad (Libreria openpyxl faltante) | usuario: %s | error: %s | fecha: %s', current_user.username, str(e), datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        return jsonify({'ok': False, 'error': 'openpyxl no instalado'}), 500

    buscar   = request.args.get('buscar', '').strip() or None
    orden    = request.args.get('orden', '')
    util_min = _parse_decimal(request.args.get('utilidad_min', ''))
    util_max = _parse_decimal(request.args.get('utilidad_max', ''))

    orden_sp_map = {
        'nombre_asc':  '', 'margen_asc':  'margen_asc',
        'margen_desc': 'margen_desc', 'costo_asc': 'costo_asc', 'costo_desc': 'costo_desc',
    }
    orden_sp = orden_sp_map.get(orden, '')

    try:
        productos = _call_sp('sp_reporte_costo_utilidad',
                             (buscar, orden_sp or None, util_min, util_max))
    except Exception as e:
        current_app.logger.error('Fallo de DB al exportar excel costo utilidad | usuario: %s | error: %s | fecha: %s', current_user.username, str(e), datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        return jsonify({'ok': False, 'error': str(e)}), 500

    # ── Estilos ──────────────────────────────────────────
    hdr_fill = PatternFill('solid', fgColor='7C4A1E')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    kpi_fill = PatternFill('solid', fgColor='F5E6D3')
    pos_font = Font(color='2D6A27', bold=True)
    neg_font = Font(color='C0522A', bold=True)
    thin     = Side(style='thin', color='D0B89A')
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center   = Alignment(horizontal='center', vertical='center')
    money    = '#,##0.00'
    pct_fmt  = '0.00"%"'

    def hcell(ws, row, col, val, w=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = hdr_font; c.fill = hdr_fill
        c.border = border; c.alignment = center
        if w:
            ws.column_dimensions[get_column_letter(col)].width = w
        return c

    def dcell(ws, row, col, val, fmt=None, fnt=None):
        c = ws.cell(row=row, column=col, value=val)
        c.border = border
        if fmt: c.number_format = fmt
        if fnt: c.font = fnt
        return c

    wb = Workbook()
    ws = wb.active
    ws.title = 'Costo y Ganancia'

    # Título
    ws.merge_cells('A1:G1')
    t = ws['A1']
    t.value = 'Costo y Ganancia por Producto — Dulce Migaja'
    t.font  = Font(bold=True, size=13, color='7C4A1E')
    t.alignment = center
    ws.row_dimensions[1].height = 28

    # Encabezados
    cols = [
        ('Producto',         28),
        ('Rendimiento',      14),
        ('Unidad',           10),
        ('Costo / Pieza',    16),
        ('Precio Venta',     16),
        ('Ganancia / Pieza', 16),
        ('Margen %',         12),
    ]
    R = 3
    for ci, (lbl, w) in enumerate(cols, 1):
        hcell(ws, R, ci, lbl, w)
    ws.row_dimensions[R].height = 18

    for p in productos:
        R += 1
        util = float(p.get('utilidad_unitaria', 0) or 0)
        fnt  = pos_font if util >= 0 else neg_font
        dcell(ws, R, 1, p.get('nombre_producto', ''))
        dcell(ws, R, 2, float(p.get('rendimiento', 0) or 0))
        dcell(ws, R, 3, str(p.get('unidad_rendimiento', '')))
        dcell(ws, R, 4, float(p.get('costo_unitario', 0) or 0), money)
        dcell(ws, R, 5, float(p.get('precio_venta', 0) or 0), money)
        dcell(ws, R, 6, util, money, fnt)
        dcell(ws, R, 7, float(p.get('margen_pct', 0) or 0), pct_fmt)
        ws.row_dimensions[R].height = 15

    ws.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f'costo_utilidad_{datetime.date.today().isoformat()}.xlsx'
    current_app.logger.info('Reporte excel Costo Utilidad generado exitosamente | usuario: %s | archivo: %s | fecha: %s', current_user.username, nombre, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nombre}"'}
    )


@costoUtilidad.route('/costo-utilidad/api/detalle/<int:id_receta>', methods=['GET'])
@login_required
@roles_required('admin')
def api_detalle_costo(id_receta):
    try:
        detalle = _call_sp('sp_detalle_costo_producto', (id_receta,))
        return jsonify({'ok': True, 'detalle': detalle})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── Utilidad por Ventas ─────────────────────────────────────────────

@costoUtilidad.route('/utilidad-diaria', methods=['GET'])
@login_required
@roles_required('admin')
def index_utilidad():
    current_app.logger.info('Vista de Utilidad Diaria accesada | usuario: %s | fecha: %s', current_user.username, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    hoy   = datetime.date.today()
    hace30 = hoy - datetime.timedelta(days=30)
    return render_template(
        'costoUtilidad/utilidad.html',
        fecha_inicio_default=hace30.isoformat(),
        fecha_fin_default=hoy.isoformat(),
    )


@costoUtilidad.route('/utilidad-diaria/api/reporte', methods=['GET'])
@login_required
@roles_required('admin')
def api_reporte_utilidad():
    """
    Devuelve JSON con 3 claves:
      kpis     → totales del período
      productos → resumen agrupado por producto
      detalle  → línea a línea de ventas
    """
    hoy    = datetime.date.today()
    fecha_ini = _parse_date(request.args.get('fecha_inicio'), hoy - datetime.timedelta(days=30))
    fecha_fin = _parse_date(request.args.get('fecha_fin'), hoy)

    # No permitir fechas futuras
    if fecha_fin > hoy:
        fecha_fin = hoy
    if fecha_ini > fecha_fin:
        fecha_ini = fecha_fin

    try:
        import json
        sets = _call_sp_multi(
            'sp_reporte_utilidad_ventas',
            (fecha_ini.isoformat(), fecha_fin.isoformat())
        )
        kpis      = sets[0][0] if sets and sets[0] else {}
        productos = sets[1]    if len(sets) > 1 else []
        detalle   = sets[2]    if len(sets) > 2 else []

        return Response(
            json.dumps({'ok': True, 'kpis': kpis, 'productos': productos, 'detalle': detalle},
                       default=_serial, ensure_ascii=False),
            mimetype='application/json'
        )
    except Exception as exc:
        import traceback, json
        return Response(
            json.dumps({'ok': False, 'error': str(exc)}, ensure_ascii=False),
            status=500, mimetype='application/json'
        )


@costoUtilidad.route('/utilidad-diaria/exportar-excel', methods=['GET'])
@login_required
@roles_required('admin')
def exportar_excel_utilidad():
    """
    Genera y descarga un archivo .xlsx con:
      Hoja 1 → Resumen por producto
      Hoja 2 → Detalle por venta
    """
    current_app.logger.info('Reporte excel Utilidad Diaria solicitado | usuario: %s | fecha: %s', current_user.username, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        current_app.logger.error('Fallo al exportar excel utilidad diaria (Libreria openpyxl faltante) | usuario: %s | error: %s | fecha: %s', current_user.username, str(e), datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        return jsonify({'ok': False, 'error': 'openpyxl no instalado. Ejecuta: pip install openpyxl'}), 500

    hoy       = datetime.date.today()
    fecha_ini = _parse_date(request.args.get('fecha_inicio'), hoy - datetime.timedelta(days=30))
    fecha_fin = _parse_date(request.args.get('fecha_fin'), hoy)
    if fecha_fin > hoy:
        fecha_fin = hoy
    if fecha_ini > fecha_fin:
        fecha_ini = fecha_fin

    try:
        sets = _call_sp_multi(
            'sp_reporte_utilidad_ventas',
            (fecha_ini.isoformat(), fecha_fin.isoformat())
        )
        kpis_row  = sets[0][0] if sets and sets[0] else {}
        productos = sets[1]    if len(sets) > 1 else []
        detalle   = sets[2]    if len(sets) > 2 else []
    except Exception as exc:
        current_app.logger.error('Fallo de DB al exportar excel utilidad diaria | usuario: %s | error: %s | fecha: %s', current_user.username, str(exc), datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        return jsonify({'ok': False, 'error': str(exc)}), 500

    # ── Estilos ────────────────────────────────────────────────
    hdr_fill  = PatternFill('solid', fgColor='7C4A1E')   # marrón
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    kpi_fill  = PatternFill('solid', fgColor='F5E6D3')
    pos_font  = Font(color='2D6A27', bold=True)
    neg_font  = Font(color='C0522A', bold=True)
    thin      = Side(style='thin', color='D0B89A')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='center')
    money_fmt = '#,##0.00'
    pct_fmt   = '0.00"%"'

    def hdr_cell(ws, row, col, value, width=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font   = hdr_font
        c.fill   = hdr_fill
        c.border = border
        c.alignment = center
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width
        return c

    def data_cell(ws, row, col, value, fmt=None, fnt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.border = border
        if fmt:
            c.number_format = fmt
        if fnt:
            c.font = fnt
        return c

    wb = Workbook()

    # ══════════════════════════════════════════════
    # HOJA 1: Resumen por producto
    # ══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Resumen por Producto'

    # Título
    ws1.merge_cells('A1:I1')
    t = ws1['A1']
    t.value = f'Rentabilidad por Producto  |  {fecha_ini.strftime("%d/%m/%Y")} – {fecha_fin.strftime("%d/%m/%Y")}'
    t.font  = Font(bold=True, size=13, color='7C4A1E')
    t.alignment = center
    ws1.row_dimensions[1].height = 28

    # KPIs en fila 2
    kpi_labels = [
        ('Ingresos', float(kpis_row.get('total_ingresos', 0) or 0), money_fmt),
        ('Costo Total', float(kpis_row.get('total_costo', 0) or 0), money_fmt),
        ('Utilidad Bruta', float(kpis_row.get('total_utilidad', 0) or 0), money_fmt),
        ('Margen Prom.', float(kpis_row.get('margen_prom', 0) or 0), pct_fmt),
        ('Ventas', int(kpis_row.get('total_ventas', 0) or 0), '#,##0'),
    ]
    for i, (lbl, val, fmt) in enumerate(kpi_labels, start=1):
        lc = ws1.cell(row=2, column=i, value=lbl)
        lc.font = Font(bold=True, size=9, color='9C7C5C')
        lc.alignment = center
        lc.fill = kpi_fill
        lc.border = border
        vc = ws1.cell(row=3, column=i, value=val)
        vc.number_format = fmt
        vc.alignment = center
        vc.font = Font(bold=True, size=11, color='7C4A1E')
        vc.fill = kpi_fill
        vc.border = border

    ws1.row_dimensions[3].height = 22

    # Encabezados tabla
    cols_prod = [
        ('Producto',           20),
        ('Piezas Vendidas',    16),
        ('Precio Prom. Venta', 18),
        ('Costo Unitario',     16),
        ('Utilidad / Pieza',   16),
        ('Margen %',           12),
        ('Utilidad Total',     16),
        ('Ingresos Totales',   16),
    ]
    R = 5
    for ci, (lbl, w) in enumerate(cols_prod, start=1):
        hdr_cell(ws1, R, ci, lbl, w)
    ws1.row_dimensions[R].height = 18

    # Datos
    for prod in productos:
        R += 1
        util_u = float(prod.get('utilidad_unitaria', 0) or 0)
        util_t = float(prod.get('utilidad_total', 0) or 0)
        fnt_u  = pos_font if util_u >= 0 else neg_font
        fnt_t  = pos_font if util_t >= 0 else neg_font

        data_cell(ws1, R, 1, prod.get('nombre_producto', ''))
        data_cell(ws1, R, 2, float(prod.get('total_piezas', 0) or 0), '#,##0.##')
        data_cell(ws1, R, 3, float(prod.get('precio_prom_venta', 0) or 0), money_fmt)
        data_cell(ws1, R, 4, float(prod.get('costo_unitario', 0) or 0), money_fmt)
        data_cell(ws1, R, 5, util_u, money_fmt, fnt_u)
        data_cell(ws1, R, 6, float(prod.get('margen_pct', 0) or 0), pct_fmt)
        data_cell(ws1, R, 7, util_t, money_fmt, fnt_t)
        data_cell(ws1, R, 8, float(prod.get('ingresos_total', 0) or 0), money_fmt)
        ws1.row_dimensions[R].height = 16

    ws1.freeze_panes = 'A6'

    # ══════════════════════════════════════════════
    # HOJA 2: Detalle por venta
    # ══════════════════════════════════════════════
    ws2 = wb.create_sheet('Detalle por Venta')

    ws2.merge_cells('A1:K1')
    t2 = ws2['A1']
    t2.value = f'Detalle de Ventas  |  {fecha_ini.strftime("%d/%m/%Y")} – {fecha_fin.strftime("%d/%m/%Y")}'
    t2.font  = Font(bold=True, size=13, color='7C4A1E')
    t2.alignment = center
    ws2.row_dimensions[1].height = 28

    cols_det = [
        ('Folio Venta',      14),
        ('Fecha',            12),
        ('Hora',             10),
        ('Producto',         24),
        ('Cantidad',         10),
        ('Precio Venta',     14),
        ('Costo Unitario',   14),
        ('Utilidad / Pieza', 14),
        ('Utilidad Total',   14),
        ('Ingreso Total',    14),
    ]
    R2 = 3
    for ci, (lbl, w) in enumerate(cols_det, start=1):
        hdr_cell(ws2, R2, ci, lbl, w)
    ws2.row_dimensions[R2].height = 18

    for det in detalle:
        R2 += 1
        util_u = float(det.get('utilidad_unitaria', 0) or 0)
        util_t = float(det.get('utilidad_total', 0) or 0)
        fnt_u  = pos_font if util_u >= 0 else neg_font
        fnt_t  = pos_font if util_t >= 0 else neg_font

        data_cell(ws2, R2, 1,  str(det.get('folio_venta', '')))
        data_cell(ws2, R2, 2,  str(det.get('fecha_venta', '')))
        data_cell(ws2, R2, 3,  str(det.get('hora_venta', '')))
        data_cell(ws2, R2, 4,  str(det.get('nombre_producto', '')))
        data_cell(ws2, R2, 5,  float(det.get('cantidad', 0) or 0), '#,##0.##')
        data_cell(ws2, R2, 6,  float(det.get('precio_venta', 0) or 0), money_fmt)
        data_cell(ws2, R2, 7,  float(det.get('costo_unitario', 0) or 0), money_fmt)
        data_cell(ws2, R2, 8,  util_u, money_fmt, fnt_u)
        data_cell(ws2, R2, 9,  util_t, money_fmt, fnt_t)
        data_cell(ws2, R2, 10, float(det.get('ingreso_total', 0) or 0), money_fmt)
        ws2.row_dimensions[R2].height = 15

    ws2.freeze_panes = 'A4'

    # ── Generar archivo en memoria ───────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f'utilidad_{fecha_ini.isoformat()}_a_{fecha_fin.isoformat()}.xlsx'
    current_app.logger.info('Reporte excel Utilidad Diaria generado exitosamente | usuario: %s | archivo: %s | fecha: %s', current_user.username, nombre, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nombre}"'}
    )
