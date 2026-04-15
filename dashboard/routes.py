import datetime
import json
from decimal import Decimal

from flask import render_template, request, current_app, Response
from flask_login import login_required, current_user

from auth import roles_required
from models import db
from . import dashboard_bp
from forms import PeriodoForm


# ── Helpers BD ─────────────────────────────────────────────────

def _call_sp(sp_name, params=()):
    conn = db.engine.raw_connection()
    try:
        cursor = conn.cursor()
        cursor.callproc(sp_name, params)
        if cursor.description:
            cols = [d[0] for d in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]
        else:
            rows = []
        cursor.close()
        return rows
    finally:
        conn.close()


def _call_sp_multi(sp_name, params=()):
    conn = db.engine.raw_connection()
    try:
        cursor = conn.cursor()
        cursor.callproc(sp_name, params)
        sets = []
        while True:
            if cursor.description:
                cols = [d[0] for d in cursor.description]
                sets.append([dict(zip(cols, r)) for r in cursor.fetchall()])
            else:
                sets.append([])
            if not cursor.nextset():
                break
        cursor.close()
        return sets
    finally:
        conn.close()


def _json(data, status=200):
    def _serial(obj):
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, (datetime.date, datetime.datetime)):
            return str(obj)
        return str(obj)
    return current_app.response_class(
        response=json.dumps(data, default=_serial, ensure_ascii=False),
        status=status,
        mimetype='application/json',
    )


def _pct(actual, anterior):
    a, b = float(actual or 0), float(anterior or 0)
    if b == 0:
        return None
    return round((a - b) / b * 100, 1)


PERIODOS_VALIDOS = ('hoy', 'semanal', 'mensual', 'anual')


# ── Ruta principal ──────────────────────────────────────────────

@dashboard_bp.route('/dashboard')
@login_required
@roles_required('admin', 'empleado', 'panadero')
def index():
    form = PeriodoForm(request.args)
    now  = datetime.datetime.now()
    return render_template('dashboard.html', form=form, now=now)


# ── API · Ventas Totales (admin + empleado) ─────────────────────
# SP: sp_dash_ventas_totales(p_periodo)
# Usa vista: vw_dash_ventas_consolidadas

@dashboard_bp.route('/dashboard/api/ventas')
@login_required
@roles_required('admin', 'empleado')
def api_ventas():
    periodo = request.args.get('periodo', 'semanal')
    if periodo not in PERIODOS_VALIDOS:
        periodo = 'semanal'
    try:
        sets    = _call_sp_multi('sp_dash_ventas_totales', (periodo,))
        resumen = sets[0][0] if sets and sets[0] else {}
        serie   = sets[1]    if len(sets) > 1     else []

        actual   = float(resumen.get('total_actual',   0) or 0)
        anterior = float(resumen.get('total_anterior', 0) or 0)

        return _json({
            'ok':               True,
            'total_actual':     actual,
            'total_anterior':   anterior,
            'pct_cambio':       _pct(actual, anterior),
            'tickets_actual':   int(resumen.get('tickets_actual',   0) or 0),
            'tickets_anterior': int(resumen.get('tickets_anterior', 0) or 0),
            'serie':            serie,
        })
    except Exception as exc:
        current_app.logger.error('dashboard.api_ventas | %s', exc)
        return _json({'ok': False, 'error': str(exc)}, 500)


# ── API · Salidas de Efectivo (solo admin) ──────────────────────
# SP: sp_dash_salidas_efectivo(p_periodo)

@dashboard_bp.route('/dashboard/api/salidas')
@login_required
@roles_required('admin')
def api_salidas():
    periodo = request.args.get('periodo', 'semanal')
    if periodo not in PERIODOS_VALIDOS:
        periodo = 'semanal'
    try:
        sets    = _call_sp_multi('sp_dash_salidas_efectivo', (periodo,))
        resumen = sets[0][0] if sets and sets[0] else {}
        cats    = sets[1]    if len(sets) > 1    else []

        actual   = float(resumen.get('total_actual',   0) or 0)
        anterior = float(resumen.get('total_anterior', 0) or 0)

        return _json({
            'ok':                   True,
            'total_actual':         actual,
            'total_anterior':       anterior,
            'pct_cambio':           _pct(actual, anterior),
            'movimientos_actual':   int(resumen.get('movimientos_actual',   0) or 0),
            'movimientos_anterior': int(resumen.get('movimientos_anterior', 0) or 0),
            'por_categoria':        cats,
        })
    except Exception as exc:
        current_app.logger.error('dashboard.api_salidas | %s', exc)
        return _json({'ok': False, 'error': str(exc)}, 500)


# ── API · Utilidad Bruta por Producto (solo admin) ──────────────
# SP: sp_dash_utilidad_por_producto()
# Usa costo promedio del último mes (detalle_compras del mes).
# Fallback a v_ultimo_costo_materia cuando no hay compras recientes.

@dashboard_bp.route('/dashboard/api/utilidad')
@login_required
@roles_required('admin')
def api_utilidad():
    try:
        rows = _call_sp('sp_dash_utilidad_por_producto')
        return _json({'ok': True, 'productos': rows})
    except Exception as exc:
        current_app.logger.error('dashboard.api_utilidad | %s', exc)
        return _json({'ok': False, 'error': str(exc)}, 500)


# ── API · Top 5 Productos (admin + empleado + panadero) ─────────
# SP: sp_dash_top_productos()
# Fijo: últimos 7 días, top 5, ventas por pieza (nuevo flujo).
# Usa vista: vw_dash_piezas_vendidas

@dashboard_bp.route('/dashboard/api/top-productos')
@login_required
@roles_required('admin', 'empleado', 'panadero')
def api_top_productos():
    try:
        rows = _call_sp('sp_dash_top_productos')
        return _json({'ok': True, 'productos': rows})
    except Exception as exc:
        current_app.logger.error('dashboard.api_top_productos | %s', exc)
        return _json({'ok': False, 'error': str(exc)}, 500)


# ── API · Stock Crítico MP (admin + empleado + panadero) ────────
# SP: sp_dash_mp_criticas()
# Usa vista: vw_dash_mp_criticas

@dashboard_bp.route('/dashboard/api/mp-criticas')
@login_required
@roles_required('admin', 'empleado', 'panadero')
def api_mp_criticas():
    try:
        rows = _call_sp('sp_dash_mp_criticas')
        return _json({
            'ok':      True,
            'items':   rows,
            'total':   len(rows),
            'criticas': sum(1 for r in rows if r.get('nivel') == 'critico'),
        })
    except Exception as exc:
        current_app.logger.error('dashboard.api_mp_criticas | %s', exc)
        return _json({'ok': False, 'error': str(exc)}, 500)


@dashboard_bp.route('/dashboard/exportar-excel', methods=['GET'])
@login_required
@roles_required('admin', 'empleado', 'panadero')
def exportar_excel_dashboard():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _json({'ok': False, 'error': 'openpyxl no instalado'}), 500

    periodo = request.args.get('periodo', 'semanal')
    if periodo not in PERIODOS_VALIDOS:
        periodo = 'semanal'

    rol = current_user.rol.clave_rol

    # ── Estilos base ────────────────────────────────────────────────────────
    hdr_fill  = PatternFill('solid', fgColor='7C4A1E')
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    sub_fill  = PatternFill('solid', fgColor='F5E6D3')
    sub_font  = Font(bold=True, color='7C4A1E', size=11)
    pos_font  = Font(color='2D6A27', bold=True)
    neg_font  = Font(color='C0522A', bold=True)
    thin      = Side(style='thin', color='D0B89A')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Formatos de número ajustados
    money_fmt = '"$"#,##0.00' # Añadido el signo de $
    pct_fmt   = '0.0"%"'
    int_fmt   = '#,##0'

    def format_hdr(ws, row, col, val, w=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = hdr_font; c.fill = hdr_fill
        c.border = border; c.alignment = center
        if w: ws.column_dimensions[get_column_letter(col)].width = w
        return c

    def format_cell(ws, row, col, val, fmt=None, fnt=None, align=None):
        c = ws.cell(row=row, column=col, value=val)
        c.border = border
        if fmt: c.number_format = fmt
        if fnt: c.font = fnt
        if align: c.alignment = align
        return c

    # Helper para limpiar decimales innecesarios (ej. 4.0 -> 4)
    def clean_num(val):
        try:
            v = float(val or 0)
            return int(v) if v.is_integer() else v
        except:
            return 0

    wb = Workbook()
    
    # ════════════════════════════════════════════════════════════
    # HOJA 1: RESUMEN GENERAL (Todos)
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Resumen General'

    ws1.merge_cells('A1:D1') # Ajustado a 4 columnas
    t1 = ws1['A1']
    t1.value = f'Dashboard - Resumen General ({periodo.capitalize()})'
    t1.font  = Font(bold=True, size=13, color='7C4A1E')
    t1.alignment = center
    ws1.row_dimensions[1].height = 28

    current_row = 3

    # -- Ventas y Salidas --
    if rol in ('admin', 'empleado'):
        v_sets = _call_sp_multi('sp_dash_ventas_totales', (periodo,))
        res_v  = v_sets[0][0] if v_sets and v_sets[0] else {}
        
        ws1.merge_cells(f'A{current_row}:D{current_row}')
        c = ws1.cell(row=current_row, column=1, value='Ingresos del Período')
        c.font = sub_font; c.fill = sub_fill; c.alignment = center; c.border = border
        for col in range(2, 5): ws1.cell(row=current_row, column=col).border = border
        current_row += 1
        
        format_hdr(ws1, current_row, 1, 'Total Actual', 18)
        format_hdr(ws1, current_row, 2, 'Total Anterior', 18)
        format_hdr(ws1, current_row, 3, 'Ventas Actual', 18)
        format_hdr(ws1, current_row, 4, 'Ventas Anterior', 18)
        current_row += 1
        
        format_cell(ws1, current_row, 1, clean_num(res_v.get('total_actual', 0)), money_fmt)
        format_cell(ws1, current_row, 2, clean_num(res_v.get('total_anterior', 0)), money_fmt)
        format_cell(ws1, current_row, 3, clean_num(res_v.get('tickets_actual', 0)), int_fmt, align=center)
        format_cell(ws1, current_row, 4, clean_num(res_v.get('tickets_anterior', 0)), int_fmt, align=center)
        current_row += 3

    if rol == 'admin':
        s_sets = _call_sp_multi('sp_dash_salidas_efectivo', (periodo,))
        res_s  = s_sets[0][0] if s_sets and s_sets[0] else {}
        
        ws1.merge_cells(f'A{current_row}:D{current_row}')
        c = ws1.cell(row=current_row, column=1, value='Salidas Aprobadas')
        c.font = sub_font; c.fill = sub_fill; c.alignment = center; c.border = border
        for col in range(2, 5): ws1.cell(row=current_row, column=col).border = border
        current_row += 1
        
        format_hdr(ws1, current_row, 1, 'Total Actual', 18)
        format_hdr(ws1, current_row, 2, 'Total Anterior', 18)
        format_hdr(ws1, current_row, 3, 'Mov. Actuales', 18)
        format_hdr(ws1, current_row, 4, 'Mov. Anteriores', 18)
        current_row += 1
        
        format_cell(ws1, current_row, 1, clean_num(res_s.get('total_actual', 0)), money_fmt, fnt=neg_font)
        format_cell(ws1, current_row, 2, clean_num(res_s.get('total_anterior', 0)), money_fmt)
        format_cell(ws1, current_row, 3, clean_num(res_s.get('movimientos_actual', 0)), int_fmt, align=center)
        format_cell(ws1, current_row, 4, clean_num(res_s.get('movimientos_anterior', 0)), int_fmt, align=center)
        current_row += 3

    # -- Top 5 Productos --
    top_prods = _call_sp('sp_dash_top_productos')
    ws1.merge_cells(f'A{current_row}:C{current_row}')
    c = ws1.cell(row=current_row, column=1, value='Top 5 Productos Más Vendidos (Últimos 7 días)')
    c.font = sub_font; c.fill = sub_fill; c.alignment = center; c.border = border
    for col in range(2, 4): ws1.cell(row=current_row, column=col).border = border
    current_row += 1
    
    format_hdr(ws1, current_row, 1, 'Producto', 30)
    format_hdr(ws1, current_row, 2, 'Piezas Vendidas', 20)
    format_hdr(ws1, current_row, 3, 'Ingresos Generados', 20)
    current_row += 1
    
    for p in top_prods:
        format_cell(ws1, current_row, 1, p.get('nombre_producto', ''))
        format_cell(ws1, current_row, 2, clean_num(p.get('total_piezas', 0)), int_fmt, align=center)
        format_cell(ws1, current_row, 3, clean_num(p.get('total_ingresos', 0)), money_fmt, fnt=pos_font)
        current_row += 1


    # ════════════════════════════════════════════════════════════
    # HOJA 2: UTILIDAD BRUTA (Solo Admin)
    # ════════════════════════════════════════════════════════════
    if rol == 'admin':
        ws2 = wb.create_sheet('Rentabilidad')
        ws2.merge_cells('A1:E1')
        t2 = ws2['A1']
        t2.value = 'Utilidad Bruta por Producto'
        t2.font  = Font(bold=True, size=13, color='7C4A1E')
        t2.alignment = center
        ws2.row_dimensions[1].height = 28
        
        cols_ut = [('Producto', 30), ('Precio de venta', 18), ('Costo unitario', 18), 
                   ('Utilidad / pza', 18), ('Margen %', 15)]
        
        for ci, (lbl, w) in enumerate(cols_ut, 1):
            format_hdr(ws2, 3, ci, lbl, w)
            
        utilidad_prods = _call_sp('sp_dash_utilidad_por_producto')
        r2 = 4
        for p in utilidad_prods:
            util = clean_num(p.get('utilidad_unitaria', 0))
            fnt_u = pos_font if util >= 0 else neg_font
            format_cell(ws2, r2, 1, p.get('nombre_producto', ''))
            format_cell(ws2, r2, 2, clean_num(p.get('precio_venta', 0)), money_fmt)
            format_cell(ws2, r2, 3, clean_num(p.get('costo_unitario', 0)), money_fmt)
            format_cell(ws2, r2, 4, util, money_fmt, fnt_u)
            format_cell(ws2, r2, 5, clean_num(p.get('margen_pct', 0)), pct_fmt)
            r2 += 1
        ws2.freeze_panes = 'A4'

    # ════════════════════════════════════════════════════════════
    # HOJA 3: INSUMOS CRÍTICOS (Todos)
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Stock Crítico')
    ws3.merge_cells('A1:E1')
    t3 = ws3['A1']
    t3.value = 'Materias Primas Bajo Stock Mínimo'
    t3.font  = Font(bold=True, size=13, color='7C4A1E')
    t3.alignment = center
    ws3.row_dimensions[1].height = 28
    
    cols_st = [('Materia Prima', 30), ('Categoría', 20), ('Stock Actual', 18), 
               ('Stock Mínimo', 18), ('Estado', 15)]
               
    for ci, (lbl, w) in enumerate(cols_st, 1):
        format_hdr(ws3, 3, ci, lbl, w)
        
    stock_critico = _call_sp('sp_dash_mp_criticas')
    r3 = 4
    for mp in stock_critico:
        stock_act = clean_num(mp.get('stock_actual', 0))
        stock_min = clean_num(mp.get('stock_minimo', 0))
        unidad = mp.get('unidad_base', '')

        # Se concatenan las cantidades con su unidad de medida y se alinean a la derecha para fácil lectura
        str_stock_act = f"{stock_act} {unidad}".strip()
        str_stock_min = f"{stock_min} {unidad}".strip()
        
        format_cell(ws3, r3, 1, mp.get('nombre', ''))
        format_cell(ws3, r3, 2, mp.get('categoria', ''))
        format_cell(ws3, r3, 3, str_stock_act, align=right_align)
        format_cell(ws3, r3, 4, str_stock_min, align=right_align)
        
        nivel = str(mp.get('nivel', '')).upper()
        if nivel == 'CRITICO' or nivel == 'CRÍTICO':
            fnt_estado = Font(color='C0522A', bold=True)
        elif nivel == 'BAJO':
            fnt_estado = Font(color='D97706', bold=True)
        else: 
            fnt_estado = Font(color='CA8A04', bold=True)
            
        format_cell(ws3, r3, 5, nivel, fnt=fnt_estado, align=center)
        r3 += 1
    ws3.freeze_panes = 'A4'

    # ── Generar archivo en memoria ──
    import io
    import datetime
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f'dashboard_{periodo}_{datetime.date.today().isoformat()}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nombre}"'}
    )