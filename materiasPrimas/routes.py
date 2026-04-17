import uuid as _uuid
import datetime
import io
from flask import render_template, request, redirect, url_for, flash, session, current_app, Response
from flask_login import login_required, current_user
from auth import roles_required
from sqlalchemy import text
from sqlalchemy.exc import OperationalError, IntegrityError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import db
from utils.db_roles import role_connection
from . import materias_primas_bp

POR_PAGINA = 10


def _usuario_actual():
    return session.get('id_usuario')


def _msg_error_sp(exc):
    if exc.orig and len(exc.orig.args) > 1:
        return exc.orig.args[1]
    return 'Ocurrió un error al procesar la operación.'


class _Paginacion:
    """Objeto liviano que imita la interfaz de Flask-SQLAlchemy Pagination."""
    def __init__(self, page, per_page, total):
        self.page     = page
        self.per_page = per_page
        self.total    = total
        self.pages    = max(1, -(-total // per_page))  # ceil division
        self.has_prev = page > 1
        self.has_next = page < self.pages
        self.prev_num = page - 1
        self.next_num = page + 1


@materias_primas_bp.route('/materias-primas', methods=['GET'])
@login_required
@roles_required('admin', 'empleado', 'panadero')
def index_materias_primas():
    current_app.logger.info('Vista de panel de materias primas accesada | usuario: %s | fecha: %s', current_user.username, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    buscar     = request.args.get('buscar', '').strip()
    estatus    = request.args.get('estatus', 'todos')
    nivel_stock = request.args.get('nivel_stock', '')
    pagina     = request.args.get('pagina', 1, type=int)
    if pagina < 1:
        pagina = 1

    # ── Construir cláusula WHERE sobre la vista ──────────────────────
    where_parts = []
    params = {}

    if buscar:
        where_parts.append("(nombre LIKE :buscar OR categoria LIKE :buscar)")
        params['buscar'] = f'%{buscar}%'
    if estatus in ('activo', 'inactivo'):
        where_parts.append("estatus = :estatus")
        params['estatus'] = estatus
    if nivel_stock in ('normal', 'bajo', 'critico'):
        where_parts.append("nivel_stock = :nivel_stock")
        params['nivel_stock'] = nivel_stock

    where_sql = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

    # ── Conteo, filas y estadísticas ─────────────────────────────────
    with role_connection() as conn:
        total_filtrado = conn.execute(
            text(f"SELECT COUNT(*) FROM vw_materias_primas {where_sql}"),
            params
        ).scalar()

        offset = (pagina - 1) * POR_PAGINA
        params_pag = {**params, 'limit': POR_PAGINA, 'offset': offset}
        lista = conn.execute(
            text(f"SELECT * FROM vw_materias_primas {where_sql} "
                 "ORDER BY nombre LIMIT :limit OFFSET :offset"),
            params_pag
        ).fetchall()

        stats = conn.execute(
            text("""
                SELECT
                    COUNT(*)                                          AS total,
                    SUM(estatus = 'activo')                           AS total_activos,
                    SUM(estatus = 'inactivo')                         AS total_inactivos,
                    SUM(nivel_stock = 'normal')                       AS stat_normal,
                    SUM(nivel_stock = 'bajo')                         AS stat_bajo,
                    SUM(nivel_stock = 'critico')                      AS stat_critico
                FROM vw_materias_primas
            """)
        ).fetchone()

    paginacion = _Paginacion(pagina, POR_PAGINA, total_filtrado)

    return render_template(
        'materiasPrimas/materiasPrimas.html',
        materias=lista,
        paginacion=paginacion,
        pagina=pagina,
        total=stats.total or 0,
        total_activos=stats.total_activos or 0,
        total_inactivos=stats.total_inactivos or 0,
        stat_normal=stats.stat_normal or 0,
        stat_bajo=stats.stat_bajo or 0,
        stat_critico=stats.stat_critico or 0,
        buscar=buscar,
        estatus_sel=estatus,
        nivel_stock_sel=nivel_stock,
    )


@materias_primas_bp.route('/materias-primas/nueva', methods=['POST'])
@login_required
@roles_required('admin', 'empleado')
def materias_primas_nueva():
    nombre      = request.form.get('nombre', '').strip()
    categoria   = request.form.get('categoria', '').strip() or None
    unidad_base = request.form.get('unidad_base', '').strip()
    stock_min   = request.form.get('stock_minimo', '0').strip()
    estatus     = request.form.get('estatus', 'activo')

    # Validar stock numérico antes de llamar al SP
    try:
        stock_min_f = float(stock_min) if stock_min else 0.0
    except ValueError:
        current_app.logger.warning(
            'Creacion de materia prima fallida (stock no numerico) | usuario: %s | materia: %s | fecha: %s',
            current_user.username, nombre, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        flash('Los valores de stock deben ser numéricos.', 'error')
        return redirect(url_for('materias_primas.index_materias_primas', modal='nueva'))

    try:
        with role_connection() as conn:
            conn.execute(
                text("CALL sp_crear_materia_prima("
                     ":uuid, :nombre, :categoria, :unidad_base, "
                     ":stock_minimo, :estatus, :creado_por)"),
                {
                    'uuid':         str(_uuid.uuid4()),
                    'nombre':       nombre,
                    'categoria':    categoria,
                    'unidad_base':  unidad_base,
                    'stock_minimo': stock_min_f,
                    'estatus':      estatus if estatus in ('activo', 'inactivo') else 'activo',
                    'creado_por':   _usuario_actual(),
                }
            )
            conn.commit()
        current_app.logger.info(
            'Materia prima creada exitosamente | usuario: %s | materia: %s | fecha: %s',
            current_user.username, nombre, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        flash(f'Materia prima "{nombre}" creada correctamente.', 'success')

    except (OperationalError, IntegrityError) as e:
        current_app.logger.warning(
            'Creacion de materia prima fallida (db) | usuario: %s | materia: %s | error: %s | fecha: %s',
            current_user.username, nombre, str(e), datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        flash(_msg_error_sp(e), 'error')
        return redirect(url_for('materias_primas.index_materias_primas', modal='nueva'))
    except Exception as e:
        current_app.logger.error(
            'Error general al crear materia prima | usuario: %s | materia: %s | error: %s | fecha: %s',
            current_user.username, nombre, str(e), datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        flash('Error al guardar la materia prima.', 'error')
        return redirect(url_for('materias_primas.index_materias_primas', modal='nueva'))

    return redirect(url_for('materias_primas.index_materias_primas'))


@materias_primas_bp.route('/materias-primas/editar/<int:id_materia>', methods=['POST'])
@login_required
@roles_required('admin', 'empleado')
def materias_primas_editar(id_materia):
    nombre      = request.form.get('nombre', '').strip()
    categoria   = request.form.get('categoria', '').strip() or None
    unidad_base = request.form.get('unidad_base', '').strip()
    stock_min   = request.form.get('stock_minimo', '').strip()
    estatus     = request.form.get('estatus', '')

    # Convertir stock_minimo (puede venir vacío → None para que el SP conserve el valor actual)
    stock_min_f = None
    if stock_min:
        try:
            stock_min_f = float(stock_min)
        except ValueError:
            flash('El stock mínimo debe ser numérico.', 'error')
            return redirect(url_for('materias_primas.index_materias_primas',
                                    modal='editar', id=id_materia))

    try:
        with role_connection() as conn:
            conn.execute(
                text("CALL sp_editar_materia_prima("
                     ":id_materia, :nombre, :categoria, :unidad_base, "
                     ":stock_minimo, :estatus, :ejecutado_por)"),
                {
                    'id_materia':   id_materia,
                    'nombre':       nombre,
                    'categoria':    categoria,
                    'unidad_base':  unidad_base,
                    'stock_minimo': stock_min_f,
                    'estatus':      estatus if estatus in ('activo', 'inactivo') else None,
                    'ejecutado_por': _usuario_actual(),
                }
            )
            conn.commit()
        current_app.logger.info(
            'Materia prima actualizada exitosamente | usuario: %s | materia: %s | fecha: %s',
            current_user.username, nombre, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        flash(f'Materia prima "{nombre}" actualizada correctamente.', 'success')

    except (OperationalError, IntegrityError) as e:
        current_app.logger.warning(
            'Edicion de materia prima fallida (db) | usuario: %s | id: %s | error: %s | fecha: %s',
            current_user.username, id_materia, str(e), datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        flash(_msg_error_sp(e), 'error')
        return redirect(url_for('materias_primas.index_materias_primas',
                                modal='editar', id=id_materia))
    except Exception as e:
        current_app.logger.error(
            'Error general al editar materia prima | usuario: %s | id: %s | error: %s | fecha: %s',
            current_user.username, id_materia, str(e), datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        flash('Error al actualizar la materia prima.', 'error')
        return redirect(url_for('materias_primas.index_materias_primas',
                                modal='editar', id=id_materia))

    return redirect(url_for('materias_primas.index_materias_primas'))


@materias_primas_bp.route('/materias-primas/exportar-excel')
@login_required
@roles_required('admin', 'empleado')
def materias_primas_exportar_excel():
    buscar      = request.args.get('buscar', '').strip()
    estatus     = request.args.get('estatus', 'todos')
    nivel_stock = request.args.get('nivel_stock', '')

    where_parts = []
    params = {}
    if buscar:
        where_parts.append("(nombre LIKE :buscar OR categoria LIKE :buscar)")
        params['buscar'] = f'%{buscar}%'
    if estatus in ('activo', 'inactivo'):
        where_parts.append("estatus = :estatus")
        params['estatus'] = estatus
    if nivel_stock in ('normal', 'bajo', 'critico'):
        where_parts.append("nivel_stock = :nivel_stock")
        params['nivel_stock'] = nivel_stock

    where_sql = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

    with role_connection() as conn:
        filas = conn.execute(
            text(f"SELECT * FROM vw_materias_primas {where_sql} ORDER BY nombre"),
            params
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Materias Primas'

    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='6B4423')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    center    = Alignment(horizontal='center', vertical='center')
    thin      = Side(style='thin', color='D4B896')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    nivel_label = {'normal': 'Normal', 'bajo': 'Stock Bajo', 'critico': 'Sin Stock'}

    headers    = ['#', 'Nombre', 'Categoría', 'Unidad Base',
                  'Stock Actual', 'Stock Mínimo', 'Nivel Stock', 'Estatus']
    col_widths = [5, 30, 20, 14, 14, 14, 14, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    for i, m in enumerate(filas, 1):
        stock  = float(m.stock_actual or 0)
        minimo = float(m.stock_minimo or 0)
        row_vals = [
            i,
            m.nombre,
            m.categoria or 'Sin categoría',
            m.unidad_base,
            stock,
            minimo,
            nivel_label.get(m.nivel_stock, m.nivel_stock),
            'Activo' if m.estatus == 'activo' else 'Inactivo',
        ]
        nivel = m.nivel_stock
        row_fill = None
        if nivel == 'critico':
            row_fill = PatternFill('solid', fgColor='FDECEA')
        elif nivel == 'bajo':
            row_fill = PatternFill('solid', fgColor='FFF8E1')

        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border    = border
            cell.alignment = center if col in (1, 3, 4, 6, 7, 8) else Alignment(vertical='center')
            if row_fill:
                cell.fill = row_fill

    fecha_str = datetime.datetime.now().strftime('%Y-%m-%d')
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    current_app.logger.info(
        'Exportación Excel materias primas | usuario: %s | filas: %d | fecha: %s',
        current_user.username, len(filas), fecha_str
    )

    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=MatPrimas_DulceMigaja_{fecha_str}.xlsx'}
    )


@materias_primas_bp.route('/materias-primas/toggle/<int:id_materia>', methods=['POST'])
@login_required
@roles_required('admin', 'empleado')
def materias_primas_toggle(id_materia):
    try:
        with role_connection() as conn:
            result = conn.execute(
                text("CALL sp_toggle_materia_prima(:id_materia, :ejecutado_por)"),
                {
                    'id_materia':    id_materia,
                    'ejecutado_por': _usuario_actual(),
                }
            )
            row = result.fetchone()
            conn.commit()

        nuevo_estatus = row.nuevo_estatus if row else 'actualizado'
        nombre_mp     = row.nombre        if row else ''
        accion        = 'activada' if nuevo_estatus == 'activo' else 'desactivada'
        current_app.logger.info(
            'Estatus de materia prima cambiado | usuario: %s | materia: %s | estatus: %s | fecha: %s',
            current_user.username, nombre_mp, nuevo_estatus,
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        flash(f'Materia prima "{nombre_mp}" {accion}.', 'success')

    except (OperationalError, IntegrityError) as e:
        current_app.logger.error(
            'Error db al cambiar estatus de materia prima | usuario: %s | id: %s | error: %s | fecha: %s',
            current_user.username, id_materia, str(e), datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        flash(_msg_error_sp(e), 'error')
    except Exception as e:
        current_app.logger.error(
            'Error general al cambiar estatus de materia prima | usuario: %s | id: %s | error: %s | fecha: %s',
            current_user.username, id_materia, str(e), datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        flash('Error al cambiar el estatus.', 'error')

    return redirect(url_for('materias_primas.index_materias_primas'))
